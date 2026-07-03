import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import tempfile
import os

# ── Constantes ─────────────────────────────────────────────────────
MESES = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
         7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}

MESES_TEXTO = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
               "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
               "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12}

PRECIO_BASE_USD = 37.8

HEADERS_SEAT = [
    "Año", "Mes", "Consumo [kWh]", "Energía facturada kWh",
    "Factor Referenciación de Precios", "Dólar",
    "Precio Base Energia [USD/MWh]", "Precio Energia Actual [USD/MWh]",
    "Precio Energia Actual [$/kWh]", "Precio Energia [$]",
    "Potecia Hr Punta [kW]", "Precio Potencia [$/kW/mes]", "Precio Potencia $",
    "Transmisión Zonal [$/kWh]", "Transmisión Zonal [$]",
    "Transmisión Nacional [$/kWh]", "Transmisión Nacional [$]",
    "Cargo asociado a exenciones [$/kWh]", "Cargo asociado a exenciones [$]",
    "Peaje SS.CC", "Precio Peaje SS.CC $",
    "Cargo por Servicio Público [$/kWh]", "Cargo por Servicio Público [$]",
    "Sobrecostos MT", "Sobrecosto PD", "Compensacion PE",
    "Costo Oportunidad RH", "Sobrecosto RH",
    "Servicios Complementarios", "Reliquidación Factura Anterior",
    "Liquidacion Peaje CEN", "Pago ERNC",
    "Total Sin IVA", "IVA", "Total C/IVA",
]

HEADERS_LIMACHE_EXTRA = [
    "Energía", "HP", "Demanda Maxima", "Pot Fact Comp",
    "Cargo fijo mensual",
    "Cargo por energía $/kWh", "Cargo por energía $",
    "Cargo por dda máx pot $/kW", "Cargo por dda máx pot $",
    "Cargo por Compra de Potencia $/kW", "Cargo por Compra de Potencia $",
    "Cargo por dda máx pot HP $/kW", "Cargo por dda máx pot HP $",
    "Estabilización Tarifas [$/kWh]", "Estabilización Tarifas $",
    "Total Sin IVA", "IVA", "Total C/IVA",
]


def safe(val, default=0):
    if pd.notna(val) and isinstance(val, (int, float)):
        return val
    return default


# ── Lectura ────────────────────────────────────────────────────────
def extraer_mes_anno(df_proforma):
    texto = str(df_proforma.iloc[2, 2]).strip().lower()
    for nombre, num in MESES_TEXTO.items():
        if nombre in texto:
            anno = int(texto.split("de")[-1].strip())
            return anno, num
    raise ValueError(f"No se pudo extraer mes/año de: '{texto}'")


def extraer_servicio(row, usar_consumo_como_facturada=False):
    consumo = safe(row[32])
    ea_bd = safe(row[36], 1)              # Energía facturada BD (indexada)
    ea = int(consumo) if usar_consumo_como_facturada else ea_bd
    cargo_ea = safe(row[38])
    pot = safe(row[37])
    cargo_pot = safe(row[39])

    # Precios unitarios desde BD
    tx_z_kwh = safe(row[26])
    tx_n_kwh = safe(row[23])
    exc_kwh = safe(row[25])
    sscc_kwh = safe(row[24])
    csp_kwh = safe(row[31])

    # $/kWh de energía: derivado con ea_bd (energía indexada) como denominador
    precio_ea_clp = cargo_ea / ea_bd if ea_bd else 0

    # CSP calculado desde $/kWh × consumo con decimales (como hace el original)
    csp_calc = csp_kwh * consumo if csp_kwh else safe(row[43])

    return {
        "consumo_kwh": consumo,
        "energia_facturada": ea,
        "factor_ref": safe(row[18], 1),
        "dolar": safe(row[21]),
        "precio_base_usd": PRECIO_BASE_USD,
        "precio_energia_usd": safe(row[12]),
        "precio_energia_clp": precio_ea_clp,
        "cargo_energia": cargo_ea,
        "pot_hp_kw": pot,
        "precio_potencia_kw": cargo_pot / pot if pot else 0,
        "cargo_potencia": cargo_pot,
        "tx_zonal_kwh": tx_z_kwh,
        "tx_zonal_pesos": tx_z_kwh * consumo,
        "tx_nacional_kwh": tx_n_kwh,
        "tx_nacional_pesos": tx_n_kwh * consumo,
        "exenciones_kwh": exc_kwh,
        "exenciones_pesos": exc_kwh * consumo,
        "sscc_kwh": sscc_kwh,
        "sscc_pesos": sscc_kwh * consumo,
        "csp_kwh": csp_kwh,
        "csp_pesos": csp_calc,
        "sobrecostos_mt": safe(row[46]),
        "sobrecosto_pd": safe(row[48]),
        "compensacion_pe": safe(row[50]),
        "costo_oportunidad_rh": 0,
        "sobrecosto_rh": 0,
        "serv_complementarios": safe(row[47]),
        "reliquidacion": 0,
        "liquidacion_peaje_cen": safe(row[49]),
        "pago_ernc": 0,
    }


def parsear_peaje(row):
    energia = safe(row[31])
    hp = safe(row[34])
    dda = safe(row[32])
    pot_fc = safe(row[33])
    estab = safe(row[18])
    dda_p = safe(row[19])
    comp_p = safe(row[20])
    hp_p = safe(row[21])
    return {
        "energia": energia, "hp": hp, "dda_max": dda, "pot_fact_comp": pot_fc,
        "cargo_fijo": safe(row[17]),
        "cargo_energia_kwh": 0, "cargo_energia_pesos": 0,
        "cargo_dda_max_kw": dda_p / dda if dda else 0, "cargo_dda_max_pesos": dda_p,
        "cargo_compra_pot_kw": comp_p / pot_fc if pot_fc else 0, "cargo_compra_pot_pesos": comp_p,
        "cargo_hp_kw": hp_p / hp if hp else 0, "cargo_hp_pesos": hp_p,
        "estab_kwh": estab / energia if energia else 0, "estab_pesos": estab,
    }


def extraer_peajes_dx(df_peajes):
    """Busca la última fila con datos reales (cargo fijo > 0)."""
    for r in range(len(df_peajes) - 1, -1, -1):
        cargo = df_peajes.iloc[r, 17]
        if pd.notna(cargo) and isinstance(cargo, (int, float)) and cargo > 0:
            return parsear_peaje(df_peajes.iloc[r])
    return parsear_peaje(df_peajes.iloc[len(df_peajes) - 1])


def leer_facturacion(ruta):
    xls = pd.ExcelFile(ruta, engine="pyxlsb")
    bd = pd.read_excel(xls, sheet_name="BD Clientes", header=None)
    pq = pd.read_excel(xls, sheet_name="PROFORMA-Q", header=None)
    pl = pd.read_excel(xls, sheet_name="PROFORMA-L", header=None)
    pdx = pd.read_excel(xls, sheet_name="PeajesDx_Limache", header=None)

    anno, mes_num = extraer_mes_anno(pq)
    quilpue = extraer_servicio(bd.iloc[2], usar_consumo_como_facturada=False)
    limache = extraer_servicio(bd.iloc[3], usar_consumo_como_facturada=True)

    quilpue["reliquidacion"] = safe(pq.iloc[40, 5]) if len(pq) > 40 else 0
    limache["reliquidacion"] = safe(pl.iloc[40, 5]) if len(pl) > 40 else 0

    peaje = extraer_peajes_dx(pdx)

    return {
        "anno": anno, "mes_num": mes_num, "mes_str": MESES[mes_num],
        "quilpue": quilpue, "limache": limache,
        "peajes_dx": peaje,
    }


# ── Generar Excel de salida ────────────────────────────────────────
def valores_seat(d):
    return [
        d["anno"], d["mes"], d["consumo_kwh"], d["energia_facturada"],
        d["factor_ref"], d["dolar"], d["precio_base_usd"], d["precio_energia_usd"],
        d["precio_energia_clp"], d["cargo_energia"], d["pot_hp_kw"],
        d["precio_potencia_kw"], d["cargo_potencia"],
        d["tx_zonal_kwh"], d["tx_zonal_pesos"],
        d["tx_nacional_kwh"], d["tx_nacional_pesos"],
        d["exenciones_kwh"], d["exenciones_pesos"],
        d["sscc_kwh"], d["sscc_pesos"],
        d["csp_kwh"], d["csp_pesos"],
        d["sobrecostos_mt"], d["sobrecosto_pd"], d["compensacion_pe"],
        d["costo_oportunidad_rh"], d["sobrecosto_rh"],
        d["serv_complementarios"], d["reliquidacion"],
        d["liquidacion_peaje_cen"], d["pago_ernc"],
    ]


def valores_peaje(p):
    return [
        p["energia"], p["hp"], p["dda_max"], p["pot_fact_comp"],
        p["cargo_fijo"], p["cargo_energia_kwh"], p["cargo_energia_pesos"],
        p["cargo_dda_max_kw"], p["cargo_dda_max_pesos"],
        p["cargo_compra_pot_kw"], p["cargo_compra_pot_pesos"],
        p["cargo_hp_kw"], p["cargo_hp_pesos"],
        p["estab_kwh"], p["estab_pesos"],
    ]


def generar_excel(datos):
    from openpyxl import Workbook

    wb = Workbook()
    anno = datos["anno"]
    mes = datos["mes_str"]

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_fmt_money = '#,##0'
    num_fmt_dec = '#,##0.000'

    # ── Hoja QUILPUE ──
    ws_q = wb.active
    ws_q.title = "Fac. Ea. SEAT"
    for c, h in enumerate(HEADERS_SEAT, 1):
        cell = ws_q.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    vals_q = valores_seat({**datos["quilpue"], "anno": anno, "mes": mes})
    r = 2
    for c, v in enumerate(vals_q, 1):
        cell = ws_q.cell(row=r, column=c, value=v)
        cell.border = thin_border
        if isinstance(v, float) and abs(v) > 1000:
            cell.number_format = num_fmt_money
        elif isinstance(v, float):
            cell.number_format = num_fmt_dec

    ws_q.cell(row=r, column=33, value=f"=J{r}+M{r}+O{r}+Q{r}+S{r}+U{r}+W{r}+X{r}+Y{r}+Z{r}+AA{r}+AB{r}+AC{r}+AD{r}+AE{r}+AF{r}")
    ws_q.cell(row=r, column=34, value=f"=(AG{r}-W{r})*0.19")
    ws_q.cell(row=r, column=35, value=f"=AG{r}+AH{r}")
    for c in [33, 34, 35]:
        ws_q.cell(row=r, column=c).border = thin_border
        ws_q.cell(row=r, column=c).number_format = num_fmt_money

    # ── Hoja LIMACHE ──
    ws_l = wb.create_sheet("Fac. Ea. Limache")
    headers_lim = HEADERS_SEAT[:32] + HEADERS_LIMACHE_EXTRA
    for c, h in enumerate(headers_lim, 1):
        cell = ws_l.cell(row=1, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    vals_l = valores_seat({**datos["limache"], "anno": anno, "mes": mes})
    vals_l += valores_peaje(datos["peajes_dx"])
    r = 2
    for c, v in enumerate(vals_l, 1):
        cell = ws_l.cell(row=r, column=c, value=v)
        cell.border = thin_border
        if isinstance(v, float) and abs(v) > 1000:
            cell.number_format = num_fmt_money
        elif isinstance(v, float):
            cell.number_format = num_fmt_dec

    ws_l.cell(row=r, column=48, value=f"=J{r}+M{r}+O{r}+Q{r}+S{r}+U{r}+W{r}+X{r}+Y{r}+Z{r}+AA{r}+AB{r}+AC{r}+AD{r}+AE{r}+AF{r}+AK{r}+AM{r}+AO{r}+AQ{r}+AS{r}+AU{r}")
    ws_l.cell(row=r, column=49, value=f"=AV{r}*0.19")
    ws_l.cell(row=r, column=50, value=f"=AV{r}+AW{r}")
    for c in [48, 49, 50]:
        ws_l.cell(row=r, column=c).border = thin_border
        ws_l.cell(row=r, column=c).number_format = num_fmt_money

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Streamlit ──────────────────────────────────────────────────────
st.set_page_config(page_title="Facturación → Proyección", page_icon="⚡")
st.title("⚡ Facturación → Proyección Energía")
st.write("Sube el archivo de facturación y descarga los datos en formato Proyección.")

fact_file = st.file_uploader(
    "📄 Archivo de Facturación (.xlsb)",
    type=["xlsb"],
)

if fact_file:
    if st.button("🔄 Procesar", type="primary", use_container_width=True):
        with st.spinner("Procesando..."):
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsb", delete=False) as tmp:
                    tmp.write(fact_file.read())
                    tmp_path = tmp.name

                try:
                    datos = leer_facturacion(tmp_path)
                finally:
                    os.unlink(tmp_path)

                st.success(f"**{datos['mes_str']} {datos['anno']}** procesado")

                # Mostrar resumen
                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("QUILPUE")
                    q = datos["quilpue"]
                    st.metric("Energía", f"{q['energia_facturada']:,.0f} kWh")
                    st.metric("Potencia HP", f"{q['pot_hp_kw']:,.0f} kW")
                    st.metric("Cargo Energía", f"${q['cargo_energia']:,.0f}")

                with col2:
                    st.subheader("LIMACHE")
                    l = datos["limache"]
                    st.metric("Energía", f"{l['energia_facturada']:,.0f} kWh")
                    st.metric("Potencia HP", f"{l['pot_hp_kw']:,.0f} kW")
                    st.metric("Cargo Energía", f"${l['cargo_energia']:,.0f}")

                # Peajes Dx
                st.subheader("Peajes Distribución Limache")
                p = datos["peajes_dx"]
                pc1, pc2, pc3 = st.columns(3)
                pc1.metric("Cargo Fijo", f"${p['cargo_fijo']:,.0f}")
                pc1.metric("Dda Máx Pot", f"${p['cargo_dda_max_pesos']:,.0f}")
                pc2.metric("Compra Pot", f"${p['cargo_compra_pot_pesos']:,.0f}")
                pc2.metric("HP", f"${p['cargo_hp_pesos']:,.0f}")
                pc3.metric("Estabilización", f"${p['estab_pesos']:,.0f}")

                # Generar Excel
                resultado = generar_excel(datos)
                nombre = f"Proyeccion_{datos['mes_str']}_{datos['anno']}.xlsx"

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=resultado,
                    file_name=nombre,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error: {e}")
else:
    st.info("👆 Sube el archivo de facturación (.xlsb) para comenzar.")
