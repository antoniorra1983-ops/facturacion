"""Escribe los datos extraídos en el archivo Proyección Energía (.xlsx)."""
from openpyxl import load_workbook

from .constants import SEAT, LIM


def _encontrar_fila(ws, anno, mes_str):
    """Busca la fila del año/mes en la hoja. Retorna número de fila o None."""
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == anno and ws.cell(row=row, column=2).value == mes_str:
            return row
    return None


def _encontrar_o_crear_fila(ws, anno, mes_str):
    """Busca la fila existente o la primera fila vacía del período."""
    fila = _encontrar_fila(ws, anno, mes_str)
    if fila:
        return fila
    for row in range(2, ws.max_row + 2):
        val = ws.cell(row=row, column=1).value
        if val is None:
            return row
        if val == anno and ws.cell(row=row, column=2).value == mes_str:
            if ws.cell(row=row, column=3).value is None:
                return row
    return ws.max_row + 1


def _escribir_seat(ws, row, d):
    """Escribe datos QUILPUE en Fac. Ea. SEAT."""
    c = ws.cell
    c(row=row, column=SEAT.AÑO, value=d["anno"])
    c(row=row, column=SEAT.MES, value=d["mes"])
    c(row=row, column=SEAT.CONSUMO, value=d["consumo_kwh"])
    c(row=row, column=SEAT.EA_FACTURADA, value=d["energia_facturada"])
    c(row=row, column=SEAT.FACTOR_REF, value=d["factor_ref"])
    c(row=row, column=SEAT.DOLAR, value=d["dolar"])
    c(row=row, column=SEAT.PRECIO_BASE, value=d["precio_base_usd"])
    c(row=row, column=SEAT.PRECIO_EA_USD, value=d["precio_energia_usd"])
    c(row=row, column=SEAT.PRECIO_EA_CLP, value=d["precio_energia_clp"])
    c(row=row, column=SEAT.CARGO_EA, value=d["cargo_energia"])
    c(row=row, column=SEAT.POT_HP_KW, value=d["pot_hp_kw"])
    c(row=row, column=SEAT.PRECIO_POT_KW, value=d["precio_potencia_kw"])
    c(row=row, column=SEAT.CARGO_POT, value=d["cargo_potencia"])
    c(row=row, column=SEAT.TX_ZONAL_KWH, value=d["tx_zonal_kwh"])
    c(row=row, column=SEAT.TX_ZONAL, value=d["tx_zonal_pesos"])
    c(row=row, column=SEAT.TX_NAC_KWH, value=d["tx_nacional_kwh"])
    c(row=row, column=SEAT.TX_NAC, value=d["tx_nacional_pesos"])
    c(row=row, column=SEAT.EXENC_KWH, value=d["exenciones_kwh"])
    c(row=row, column=SEAT.EXENC, value=d["exenciones_pesos"])
    c(row=row, column=SEAT.SSCC_KWH, value=d["sscc_kwh"])
    c(row=row, column=SEAT.SSCC, value=d["sscc_pesos"])
    c(row=row, column=SEAT.CSP_KWH, value=d["csp_kwh"])
    c(row=row, column=SEAT.CSP, value=d["csp_pesos"])
    c(row=row, column=SEAT.SOBRE_MT, value=d["sobrecostos_mt"])
    c(row=row, column=SEAT.SOBRE_PD, value=d["sobrecosto_pd"])
    c(row=row, column=SEAT.COMP_PE, value=d["compensacion_pe"])
    c(row=row, column=SEAT.COSTO_RH, value=d["costo_oportunidad_rh"])
    c(row=row, column=SEAT.SOBRE_RH, value=d["sobrecosto_rh"])
    c(row=row, column=SEAT.SERV_COMP, value=d["serv_complementarios"])
    c(row=row, column=SEAT.RELIQ, value=d["reliquidacion"])
    c(row=row, column=SEAT.PEAJE_CEN, value=d["liquidacion_peaje_cen"])
    c(row=row, column=SEAT.PAGO_ERNC, value=d["pago_ernc"])

    r = row
    # Total Sin IVA = suma de todos los montos en pesos
    c(row=r, column=SEAT.TOTAL, value=(
        f"=J{r}+M{r}+O{r}+Q{r}+S{r}+U{r}+W{r}"
        f"+X{r}+Y{r}+Z{r}+AA{r}+AB{r}+AC{r}+AD{r}+AE{r}+AF{r}"
    ))
    # IVA: CSP (col W) es exento
    c(row=r, column=SEAT.IVA, value=f"=(AG{r}-W{r})*0.19")
    c(row=r, column=SEAT.TOTAL_IVA, value=f"=AG{r}+AH{r}")


def _escribir_limache(ws, row, d, peaje):
    """Escribe datos LIMACHE en Fac. Ea. Limache."""
    c = ws.cell
    c(row=row, column=1, value=d["anno"])
    c(row=row, column=2, value=d["mes"])
    c(row=row, column=3, value=d["consumo_kwh"])
    c(row=row, column=4, value=d["energia_facturada"])
    c(row=row, column=5, value=d["factor_ref"])
    c(row=row, column=6, value=d["dolar"])
    c(row=row, column=7, value=d["precio_base_usd"])
    c(row=row, column=8, value=d["precio_energia_usd"])
    c(row=row, column=9, value=d["precio_energia_clp"])
    c(row=row, column=10, value=d["cargo_energia"])
    c(row=row, column=11, value=d["pot_hp_kw"])
    c(row=row, column=12, value=d["precio_potencia_kw"])
    c(row=row, column=13, value=d["cargo_potencia"])
    c(row=row, column=14, value=d["tx_zonal_kwh"])
    c(row=row, column=15, value=d["tx_zonal_pesos"])
    c(row=row, column=16, value=d["tx_nacional_kwh"])
    c(row=row, column=17, value=d["tx_nacional_pesos"])
    c(row=row, column=18, value=d["exenciones_kwh"])
    c(row=row, column=19, value=d["exenciones_pesos"])
    c(row=row, column=20, value=d["sscc_kwh"])
    c(row=row, column=21, value=d["sscc_pesos"])
    c(row=row, column=22, value=d["csp_kwh"])
    c(row=row, column=23, value=d["csp_pesos"])
    c(row=row, column=24, value=d["sobrecostos_mt"])
    c(row=row, column=25, value=d["sobrecosto_pd"])
    c(row=row, column=26, value=d["compensacion_pe"])
    c(row=row, column=27, value=d["costo_oportunidad_rh"])
    c(row=row, column=28, value=d["sobrecosto_rh"])
    c(row=row, column=29, value=d["serv_complementarios"])
    c(row=row, column=30, value=d["reliquidacion"])
    c(row=row, column=31, value=d["liquidacion_peaje_cen"])
    c(row=row, column=32, value=d["pago_ernc"])

    # Peajes distribución
    c(row=row, column=LIM.ENERGIA_DX, value=peaje["energia"])
    c(row=row, column=LIM.HP_DX, value=peaje["hp"])
    c(row=row, column=LIM.DDA_MAX_DX, value=peaje["dda_max"])
    c(row=row, column=LIM.POT_FACT_COMP, value=peaje["pot_fact_comp"])
    c(row=row, column=LIM.CARGO_FIJO, value=peaje["cargo_fijo"])
    c(row=row, column=LIM.CARGO_EA_KWH, value=peaje["cargo_energia_kwh"])
    c(row=row, column=LIM.CARGO_EA, value=peaje["cargo_energia_pesos"])
    c(row=row, column=LIM.CARGO_DDA_KW, value=peaje["cargo_dda_max_kw"])
    c(row=row, column=LIM.CARGO_DDA, value=peaje["cargo_dda_max_pesos"])
    c(row=row, column=LIM.CARGO_COMPRA_KW, value=peaje["cargo_compra_pot_kw"])
    c(row=row, column=LIM.CARGO_COMPRA, value=peaje["cargo_compra_pot_pesos"])
    c(row=row, column=LIM.CARGO_HP_KW, value=peaje["cargo_hp_kw"])
    c(row=row, column=LIM.CARGO_HP, value=peaje["cargo_hp_pesos"])
    c(row=row, column=LIM.ESTAB_KWH, value=peaje["estab_kwh"])
    c(row=row, column=LIM.ESTAB, value=peaje["estab_pesos"])

    r = row
    # Total: contrato + peajes Dx (solo columnas de montos $)
    c(row=r, column=LIM.TOTAL, value=(
        f"=J{r}+M{r}+O{r}+Q{r}+S{r}+U{r}+W{r}"
        f"+X{r}+Y{r}+Z{r}+AA{r}+AB{r}+AC{r}+AD{r}+AE{r}+AF{r}"
        f"+AK{r}+AM{r}+AO{r}+AQ{r}+AS{r}+AU{r}"
    ))
    c(row=r, column=LIM.IVA, value=f"=AV{r}*0.19")
    c(row=r, column=LIM.TOTAL_IVA, value=f"=AV{r}+AW{r}")


def escribir_proyeccion(ruta_proy: str, datos: dict, salida: str) -> dict:
    """
    Escribe los datos extraídos en el archivo de Proyección Energía.

    Args:
        ruta_proy: ruta al archivo .xlsx de proyección (plantilla)
        datos: dict retornado por leer_facturacion()
        salida: ruta del archivo de salida

    Returns:
        dict con clave 'mensajes' (lista de strings con el resultado)
    """
    wb = load_workbook(ruta_proy)
    mensajes = []
    anno = datos["anno"]
    mes = datos["mes_str"]

    quilpue = {**datos["quilpue"], "anno": anno, "mes": mes}
    limache = {**datos["limache"], "anno": anno, "mes": mes}

    # QUILPUE → Fac. Ea. SEAT
    ws_seat = wb["Fac. Ea. SEAT"]
    fila_q = _encontrar_o_crear_fila(ws_seat, anno, mes)
    _escribir_seat(ws_seat, fila_q, quilpue)
    mensajes.append(f"✓ QUILPUE escrito en 'Fac. Ea. SEAT' fila {fila_q}")

    # LIMACHE → Fac. Ea. Limache
    ws_lim = wb["Fac. Ea. Limache"]
    fila_l = _encontrar_o_crear_fila(ws_lim, anno, mes)
    _escribir_limache(ws_lim, fila_l, limache, datos["peajes_dx"])
    mensajes.append(f"✓ LIMACHE escrito en 'Fac. Ea. Limache' fila {fila_l}")

    wb.save(salida)
    return {"mensajes": mensajes}
